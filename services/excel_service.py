import io
import openpyxl
from openpyxl.styles import Font
from utils.logger import logger

def generate_excel_report(summary, orders, expenses, income):
    """Generates an in-memory .xlsx workbook."""
    try:
        wb = openpyxl.Workbook()

        # Worksheet 1: Summary
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary.append(["Metric", "Value"])
        ws_summary.append(["Period Start", summary['start_date']])
        ws_summary.append(["Period End", summary['end_date']])
        ws_summary.append(["Orders", summary['orders_count']])
        ws_summary.append(["Sales", summary['sales']])
        ws_summary.append(["Expenses", summary['expenses']])
        ws_summary.append(["Income", summary['income']])
        ws_summary.append(["Profit", summary['profit']])

        # Worksheet 2: Orders
        ws_orders = wb.create_sheet(title="Orders")
        ws_orders.append(["Order Number", "Customer", "Phone Number", "Address", "Product", "Quantity", "Unit Price", "Delivery Fee", "Total", "Delivery Date", "Created By", "Created At"])
        for o in orders:
            product_val = o.get('product_name') or f"ID: {o.get('product_id')}"
            created_at = o['created_at'].strftime("%Y-%m-%d %H:%M:%S") if o.get('created_at') else ""
            ws_orders.append([o['order_number'], o['customer_name'], o['phone'], o['address'], product_val, o['quantity'], float(o['unit_price']), float(o['delivery_fee']), float(o['total_price']), o['delivery_date'], o['created_by'], created_at])

        # Worksheet 3: Expenses
        ws_exp = wb.create_sheet(title="Expenses")
        ws_exp.append(["Amount", "Description", "Created By", "Created At"])
        for e in expenses:
            created_at = e['created_at'].strftime("%Y-%m-%d %H:%M:%S") if e.get('created_at') else ""
            ws_exp.append([float(e['amount']), e['description'], e['created_by'], created_at])

        # Worksheet 4: Income
        ws_inc = wb.create_sheet(title="Income")
        ws_inc.append(["Amount", "Description", "Created By", "Created At"])
        for i in income:
            created_at = i['created_at'].strftime("%Y-%m-%d %H:%M:%S") if i.get('created_at') else ""
            ws_inc.append([float(i['amount']), i['description'], i['created_by'], created_at])

        # Formatting Pass
        currency_format = '"$"#,##0.00'
        
        for ws in wb.worksheets:
            # Bold headers & Freeze top row
            for cell in ws[1]:
                cell.font = Font(bold=True)
            ws.freeze_panes = 'A2'
            
            # Auto-size columns based on content
            for col in ws.columns:
                max_length = 0
                column_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = max_length + 2

        # Apply currency format to specific cells
        for row in range(5, 10):
            ws_summary[f'B{row}'].number_format = currency_format
            
        for row in range(2, len(orders) + 2):
            ws_orders[f'G{row}'].number_format = currency_format
            ws_orders[f'H{row}'].number_format = currency_format
            ws_orders[f'I{row}'].number_format = currency_format
            
        for row in range(2, len(expenses) + 2):
            ws_exp[f'A{row}'].number_format = currency_format
            
        for row in range(2, len(income) + 2):
            ws_inc[f'A{row}'].number_format = currency_format

        # Save to memory stream to send via Telegram
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        return file_stream
        
    except Exception as e:
        logger.error(f"Excel Generation Error: {e}")
        raise e